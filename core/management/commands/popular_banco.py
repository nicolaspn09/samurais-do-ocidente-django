# core/management/commands/popular_banco.py
import os
from datetime import datetime
from django.core.management.base import BaseCommand
from django.conf import settings
from core.models import Atleta, Academia, Faixa, Modalidade, HistoricoGraduacao

try:
    import openpyxl
except ImportError:
    openpyxl = None

class Command(BaseCommand):
    help = 'Popula o banco de dados lendo as abas de um único arquivo XLSX'

    def handle(self, *args, **kwargs):
        if not openpyxl:
            self.stdout.write(self.style.ERROR("A biblioteca 'openpyxl' não está instalada. Rode: pip install openpyxl"))
            return

        # Ajuste o nome do arquivo para o nome exato que você salvar na pasta
        caminho_arquivo = os.path.join(settings.BASE_DIR, 'dados_csv', 'planilha_graduacao.xlsx')
        
        if not os.path.exists(caminho_arquivo):
            self.stdout.write(self.style.ERROR(f"Arquivo não encontrado: {caminho_arquivo}"))
            return

        modalidade, _ = Modalidade.objects.get_or_create(nome='Kickboxing', defaults={'descricao': 'Importado do Excel'})
        academia_padrao, _ = Academia.objects.get_or_create(nome='Sem Academia/Desconhecida', defaults={'endereco': 'Não informado'})

        # Mapeamento do nome exato da aba (guia) na planilha para a Faixa no banco
        abas_esperadas = [
            ('Faixa-Laranja', 'Laranja', 1),
            ('Faixa-Azul', 'Azul', 2),
            ('Faixa-Amarela', 'Amarela', 3),
            ('Faixa-Verde', 'Verde', 4),
            ('Faixa-Marrom', 'Marrom', 5),
            ('Faixa-Preta', 'Preta 1º Dan', 6), # Trataremos a preta como 1º Dan no MVP
        ]

        self.stdout.write("Lendo arquivo XLSX... Isso pode levar alguns segundos.")
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)

        for nome_aba, nome_faixa, ordem in abas_esperadas:
            if nome_aba not in wb.sheetnames:
                self.stdout.write(self.style.WARNING(f"Aba '{nome_aba}' não encontrada na planilha. Pulando..."))
                continue

            faixa_obj, _ = Faixa.objects.get_or_create(nome=nome_faixa, defaults={'ordem': ordem})
            ws = wb[nome_aba]

            # Encontra os índices das colunas (Lê a linha 2, pois a linha 1 é o título mesclado)
            cabecalhos = [str(c.value).strip().lower() if c.value else '' for c in ws[2]]
            
            if 'nome' not in cabecalhos:
                self.stdout.write(self.style.ERROR(f"Aba '{nome_aba}' não tem coluna 'Nome'. Pulando..."))
                continue

            idx_nome = cabecalhos.index('nome')
            
            # Tratamento da inconsistência da coluna de Data (Preta vs Coloridas)
            idx_data = -1
            if 'data' in cabecalhos:
                idx_data = cabecalhos.index('data')
            elif '1° dan' in cabecalhos:
                idx_data = cabecalhos.index('1° dan')

            # Tratamento da inconsistência da coluna de Academia
            idx_academia = -1
            if 'academia' in cabecalhos:
                idx_academia = cabecalhos.index('academia')
            elif 'academias' in cabecalhos:
                idx_academia = cabecalhos.index('academias')

            # Lê os dados a partir da linha 3
            for row in ws.iter_rows(min_row=3, values_only=True):
                nome_Atleta = str(row[idx_nome]).strip() if row[idx_nome] else ''
                
                if not nome_Atleta or nome_Atleta.lower() == 'none':
                    continue # Pula linhas vazias

                # Resolve a academia
                nome_academia = str(row[idx_academia]).strip().upper() if idx_academia != -1 and row[idx_academia] else ''
                if nome_academia and nome_academia.lower() != 'none':
                    academia_obj, _ = Academia.objects.get_or_create(
                        nome=nome_academia,
                        defaults={'endereco': 'Endereço não cadastrado no XLSX'}
                    )
                else:
                    academia_obj = academia_padrao

                # Resolve a Data (o openpyxl geralmente já retorna um objeto datetime)
                data_graduacao = None
                valor_data = row[idx_data] if idx_data != -1 else None

                if isinstance(valor_data, datetime):
                    data_graduacao = valor_data.date()
                elif isinstance(valor_data, str) and valor_data.strip():
                    try:
                        # Tenta converter caso o Excel tenha salvado como texto
                        data_graduacao = datetime.strptime(valor_data.strip(), '%Y-%m-%d').date()
                    except ValueError:
                        pass
                
                if not data_graduacao:
                    data_graduacao = datetime.today().date()

                # Salva no Banco Relacional
                Atleta_obj, created = Atleta.objects.get_or_create(
                    nome=nome_Atleta,
                    defaults={'academia': academia_obj, 'faixa_atual': faixa_obj}
                )
                Atleta_obj.modalidades.add(modalidade)

                # Atualiza a faixa caso seja uma promoção
                if Atleta_obj.faixa_atual.ordem < faixa_obj.ordem:
                    Atleta_obj.faixa_atual = faixa_obj
                    Atleta_obj.save()

                # Registra o Histórico
                HistoricoGraduacao.objects.get_or_create(
                    atleta=Atleta_obj,
                    faixa=faixa_obj,
                    defaults={'data_graduacao': data_graduacao, 'examinador': 'Importado via XLSX'}
                )

            self.stdout.write(self.style.SUCCESS(f"Aba '{nome_aba}' importada com sucesso!"))

        self.stdout.write(self.style.SUCCESS("Migração do arquivo XLSX finalizada com sucesso."))